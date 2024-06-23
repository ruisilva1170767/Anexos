import openpyxl
import os
import cv2
from skimage.feature import local_binary_pattern
import numpy as np

dir_path = 'C:/Users/ruipsilv/OneDrive - Capgemini/Desktop/PESTA/Camera_Dataset/'
wb = openpyxl.load_workbook('C:/Users/ruipsilv/OneDrive - Capgemini/Desktop/Dataset Docs/pixel_annotations_bbox.xlsx')
ws = wb['Annotations']

def calcular_histograma_lbp(imagem, num_points, raio, metodo='uniform'):
    # Aplicar LBP
    lbp = local_binary_pattern(imagem, num_points, raio, method=metodo)

    # Construir o histograma de LBP
    n_bins = int(lbp.max() + 1)
    hist, _ = np.histogram(lbp, bins=n_bins, range=(0, n_bins), density=True)

    return hist

# Definir os parâmetros do LBP
num_points = 64  # Número de pontos ao redor do pixel central
raio = 0.2  # Raio do círculo

#Definir as bins do histograma de cores
bins = 64

for i in range(2, ws.max_row + 1):
    #Imprimir os números. Programa acaba em 101*20 = 2020
	print(i)
    #Adquirir posição da app na foto
	xi, yi, xf, yf = ws.cell(i, 3).value, ws.cell(i, 5).value, ws.cell(i, 4).value, ws.cell(i, 6).value

    #Adquirir imagem da app respetiva
	path = str(ws.cell(i, 1).value)
	img = cv2.imread(os.path.join(dir_path, path))
	hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
	grey = cv2.cvtColor(img, cv2.IMREAD_GRAYSCALE)
	hsv = hsv[yi:yf, xi:xf]
	grey = grey[yi:yf, xi:xf]
    
    #Adquirir histograma de cores de cada canal de cada app
	H, S, V = cv2.calcHist([hsv], [0], None, [bins], [0, 256]), cv2.calcHist([hsv], [1], None, [bins], [0, 256]), cv2.calcHist([hsv], [2], None, [bins], [0, 256])
    
	#Adquirir histograma LBP de cada app
	histograma_lbp = calcular_histograma_lbp(grey, num_points, raio)
    
	for j in range(0, bins):
		ws.cell(i, 7 + j, value=int(H[j, 0]))
		ws.cell(i, 7 + bins + j, value=int(S[j, 0]))
		ws.cell(i, 7 + 2*bins + j, value=int(V[j, 0]))
    
	for j in range(len(histograma_lbp)):
		ws.cell(i, 7 + 3*bins + j, value=float(histograma_lbp[j]))

wb.save('C:/Users/ruipsilv/OneDrive - Capgemini/Desktop/Dataset Docs/pixel_annotations_bbox.xlsx')